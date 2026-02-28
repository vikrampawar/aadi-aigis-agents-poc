"""
Excel ingestion for Agent 02 — VDR Financial & Operational Data Store.

Two-pass read strategy:
  Pass 1 (data_only=False) — reads formula strings
  Pass 2 (data_only=True)  — reads last-computed cached values

Stores every non-empty cell in excel_cells, plus routes typed rows
to production_series / financial_series / etc.
"""

from __future__ import annotations

import re
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from aigis_agents.agent_02_data_store import db_manager as db


# ── Constants ──────────────────────────────────────────────────────────────────

# Extensions handled by this module
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}

# Max rows/cols to scan per sheet (safety cap)
MAX_ROWS = 5_000
MAX_COLS = 500

# Threshold for detecting header rows (% of non-numeric cells in a row)
HEADER_ROW_ALPHA_THRESHOLD = 0.6


# ── Public API ─────────────────────────────────────────────────────────────────

def ingest_excel(
    file_path: str | Path,
    deal_id: str,
    doc_id: str,
    conn: sqlite3.Connection,
    case_name: str | None = None,
    sheet_names: list[str] | None = None,
    classify_fn: Any | None = None,   # semantic_classifier.classify_sheet
) -> dict[str, Any]:
    """
    Ingest an Excel workbook into the DB.

    Args:
        file_path:    Path to the .xlsx/.xlsm file.
        deal_id:      Deal UUID.
        doc_id:       Source document UUID (pre-inserted in source_documents).
        conn:         Open SQLite connection.
        case_name:    Default case tag for all cells.
        sheet_names:  Limit to these sheets (None = all).
        classify_fn:  Optional callable(sheet_name, headers, sample) → SheetClassification.

    Returns dict with: sheets_ingested, cells_written, formula_cells, assumption_cells,
                       typed_rows_routed, circular_refs_detected
    """
    path = Path(file_path)

    # ── Two-pass read ─────────────────────────────────────────────────────────
    wb_formulas = load_workbook(str(path), data_only=False, read_only=True)
    wb_values   = load_workbook(str(path), data_only=True,  read_only=True)

    stats = {
        "sheets_ingested":       0,
        "cells_written":         0,
        "formula_cells":         0,
        "assumption_cells":      0,
        "typed_rows_routed":     0,
        "circular_refs_detected": 0,
    }

    target_sheets = set(sheet_names) if sheet_names else None

    for sheet_name in wb_formulas.sheetnames:
        if target_sheets and sheet_name not in target_sheets:
            continue

        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name] if sheet_name in wb_values.sheetnames else None

        sheet_stats = _ingest_sheet(
            ws_formulas=ws_f,
            ws_values=ws_v,
            deal_id=deal_id,
            doc_id=doc_id,
            conn=conn,
            sheet_name=sheet_name,
            sheet_index=wb_formulas.sheetnames.index(sheet_name),
            default_case=case_name,
            classify_fn=classify_fn,
        )

        for k, v in sheet_stats.items():
            stats[k] = stats.get(k, 0) + v
        stats["sheets_ingested"] += 1

    wb_formulas.close()
    wb_values.close()

    return stats


# ── Sheet-level ingestion ──────────────────────────────────────────────────────

def _ingest_sheet(
    ws_formulas, ws_values,
    deal_id: str,
    doc_id: str,
    conn: sqlite3.Connection,
    sheet_name: str,
    sheet_index: int,
    default_case: str | None,
    classify_fn: Any | None,
) -> dict[str, int]:
    """Process a single sheet and write cells + sheet metadata to DB."""

    stats = {"cells_written": 0, "formula_cells": 0,
             "assumption_cells": 0, "typed_rows_routed": 0,
             "circular_refs_detected": 0}

    # Build value cache from pass 2
    value_cache: dict[str, Any] = {}
    if ws_values:
        for row in ws_values.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS):
            for cell in row:
                value_cache[cell.coordinate] = cell.value

    # Detect headers (first non-empty rows with mostly text)
    row_headers: dict[int, str] = {}     # row_num → row label
    col_headers: dict[int, str] = {}     # col_num → col label (from first text row)
    _detect_headers(ws_formulas, row_headers, col_headers)

    # Classify sheet type if classifier provided
    sheet_type_str: str | None = None
    if classify_fn:
        try:
            sample_headers = list(col_headers.values())[:10]
            sample_rows = _get_sample_rows(ws_formulas, n=3)
            classification = classify_fn(sheet_name, sample_headers, sample_rows)
            sheet_type_str = classification.sheet_type.value if classification else None
        except Exception:
            pass

    # Collect cells for bulk insert
    cell_rows: list[dict[str, Any]] = []
    formula_count = 0
    assumption_count = 0
    circular_count = 0

    for row in ws_formulas.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS):
        for cell in row:
            addr = cell.coordinate
            formula_val = cell.value
            cached_val  = value_cache.get(addr)

            # Skip truly empty cells
            if formula_val is None and cached_val is None:
                continue

            is_formula = isinstance(formula_val, str) and formula_val.startswith("=")
            has_circular = False

            # Numeric value: prefer cached for formulas, literal otherwise
            numeric_value: float | None = None
            if is_formula:
                formula_count += 1
                if isinstance(cached_val, (int, float)) and not isinstance(cached_val, bool):
                    numeric_value = float(cached_val)
                # Detect circular reference (cached value same as formula string)
                if isinstance(cached_val, str) and "#" in cached_val:
                    has_circular = True
                    circular_count += 1
            else:
                if isinstance(formula_val, (int, float)) and not isinstance(formula_val, bool):
                    numeric_value = float(formula_val)

            # Is this an assumption cell?
            is_assumption = (not is_formula) and (numeric_value is not None)
            if is_assumption:
                assumption_count += 1

            # Data type label
            if has_circular:
                data_type = "error"
            elif is_formula:
                data_type = "formula"
            elif isinstance(formula_val, bool):
                data_type = "boolean"
            elif isinstance(formula_val, (int, float)):
                data_type = "numeric"
            elif hasattr(formula_val, "date"):  # datetime
                data_type = "date"
            elif formula_val is None:
                data_type = "empty"
            else:
                data_type = "text"

            row_hdr = row_headers.get(cell.row, "")
            col_hdr = col_headers.get(cell.column, "")

            cell_rows.append({
                "deal_id":           deal_id,
                "doc_id":            doc_id,
                "sheet_name":        sheet_name,
                "cell_address":      addr,
                "row_num":           cell.row,
                "col_num":           cell.column,
                "raw_value":         str(formula_val) if formula_val is not None else None,
                "numeric_value":     numeric_value,
                "formula":           formula_val if is_formula else None,
                "data_type":         data_type,
                "number_format":     getattr(cell, "number_format", None),
                "row_header":        row_hdr or None,
                "col_header":        col_hdr or None,
                "is_assumption":     1 if is_assumption else 0,
                "is_output":         1 if (is_formula and numeric_value is not None) else 0,
                "case_name":         default_case,
            })

    # Bulk insert cells
    written = db.bulk_insert_excel_cells(conn, cell_rows)
    stats["cells_written"]         = written
    stats["formula_cells"]         = formula_count
    stats["assumption_cells"]      = assumption_count
    stats["circular_refs_detected"] = circular_count

    # Write sheet metadata
    ingest_notes = None
    if circular_count > 0:
        ingest_notes = f"xlcalculator_unsupported: {circular_count} circular reference(s) detected"

    db.insert_excel_sheet(conn, {
        "deal_id":          deal_id,
        "doc_id":           doc_id,
        "sheet_name":       sheet_name,
        "sheet_index":      sheet_index,
        "sheet_type":       sheet_type_str,
        "row_count":        ws_formulas.max_row,
        "col_count":        ws_formulas.max_column,
        "assumption_cells": assumption_count,
        "output_cells":     sum(1 for r in cell_rows if r["is_output"]),
        "formula_cells":    formula_count,
        "ingest_notes":     ingest_notes,
    })

    return stats


# ── Header detection ───────────────────────────────────────────────────────────

def _detect_headers(
    ws,
    row_headers: dict[int, str],
    col_headers: dict[int, str],
) -> None:
    """
    Populate row_headers and col_headers dicts from the first text rows/columns.
    Heuristic: a row is a header if >=60% of its non-empty cells contain text.
    """
    # Sample first 20 rows and 3 columns to find col headers
    col_header_found = False
    for row_idx, row in enumerate(ws.iter_rows(max_row=min(20, MAX_ROWS), max_col=MAX_COLS), start=1):
        non_empty = [c for c in row if c.value is not None]
        if not non_empty:
            continue
        text_cells = [c for c in non_empty
                      if isinstance(c.value, str) and not _is_numeric_string(c.value)]
        alpha_ratio = len(text_cells) / len(non_empty) if non_empty else 0

        if alpha_ratio >= HEADER_ROW_ALPHA_THRESHOLD and not col_header_found:
            # This row is a column header row
            for cell in row:
                if cell.value is not None:
                    col_headers[cell.column] = str(cell.value).strip()
            col_header_found = True
        elif text_cells and len(text_cells) <= 2:
            # Leftmost text cell is likely a row label
            for cell in row:
                if isinstance(cell.value, str) and not _is_numeric_string(cell.value):
                    row_headers[cell.row] = str(cell.value).strip()
                    break


def _is_numeric_string(s: str) -> bool:
    try:
        float(s.replace(",", "").replace("%", "").replace("$", ""))
        return True
    except ValueError:
        return False


def _get_sample_rows(ws, n: int = 3) -> list[list[str]]:
    """Return first n data rows as list of string lists."""
    result = []
    for row in ws.iter_rows(max_row=min(10, MAX_ROWS), max_col=min(20, MAX_COLS)):
        row_vals = [str(c.value) if c.value is not None else "" for c in row]
        if any(v for v in row_vals):
            result.append(row_vals)
        if len(result) >= n:
            break
    return result
